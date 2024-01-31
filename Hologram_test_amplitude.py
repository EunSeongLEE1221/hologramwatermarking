import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) 
warnings.filterwarnings("ignore", category=RuntimeWarning) 
warnings.filterwarnings("ignore", category=FutureWarning) 
import numpy as np
import tensorflow as tf
from functions import *
from attack_h import *
#from utils import *
import os
import scipy

import time
import random
import h5py
os.environ["CUDA_DEVICE_ORDER"]="PCI_BUS_ID"   
os.environ["CUDA_VISIBLE_DEVICES"]="1"
import warnings
warnings.filterwarnings("ignore", category=RuntimeWarning) 
import matplotlib.pylab as plt
import openpyxl 
from openpyxl import Workbook
import json


class WM_NN(object):
    def __init__(self, width=128, height=128, kernel1=3, kernel2=7, batch_size=32, d1_lr=0.0001, g1_lr=0.000025, gan2_lr=0.0001, beta1=0.5, beta2=0.9, lambda1=0.1, lambda2=0.1, lambda3=1.0, alpha=1.0,in_maxmin=1,in_min=-1):
        self.width = width
        self.height = height
        self.x_dim = self.width * self.height
        self.batch_size = batch_size
        self.d1_lr = d1_lr
        self.g1_lr = g1_lr
        self.gan2_lr = gan2_lr
        self.beta1 = beta1
        self.beta2 = beta2
        self.kernel1 = kernel1
        self.kernel2 = kernel2
        self.lambda1 = lambda1
        self.lambda2 = lambda2
        self.lambda3 = lambda3
        self.f = 0.0
        self.r = 1.0
        self.alpha = alpha
        self.in_maxmin = in_maxmin
        self.in_min = in_min
    def generator_1_up(self, z1, is_train=True, reuse=tf.AUTO_REUSE):
        if not is_train :
         reuse = tf.AUTO_REUSE
        with tf.variable_scope("gen1_up", reuse=tf.AUTO_REUSE):
            z = z1
            # generator network(using InfoGAN)
            net = {}
            # 1st de-conv. layer inp : (batch_size, width/16, height/16, 1)
            net[1] = deconv2D(z, [self.batch_size, round(self.height/8), round(self.width/8), 512], self.kernel1, self.kernel1, 2, 2, name='g1_up_deconv1')
            net[1] = batch_norm(net[1], is_train=is_train, name='g1_up_bnorm1')
            net[1] = tf.nn.relu(net[1])
            net[1] = tf.nn.avg_pool(net[1], ksize=[1, 2, 2, 1], strides=[1, 1, 1, 1], padding='SAME', name='g1_up_avg1')
            # 2th de-conv. layer inp : (batch_size, width/8, height/8, 128)
            net[2] = deconv2D(net[1], [self.batch_size, round(self.height/4), round(self.width/4), 256], self.kernel1, self.kernel1, 2, 2, name='g1_up_deconv2')
            net[2] = batch_norm(net[2], is_train=is_train, name='g1_up_bnorm2')
            net[2] = tf.nn.relu(net[2])
            net[2] = tf.nn.avg_pool(net[2], ksize=[1, 2, 2, 1], strides=[1, 1, 1, 1], padding='SAME', name='g1_up_avg2')
            # 3th de-conv. layer inp : (batch_size, width/4, height/4, 64)
            net[3] = deconv2D(net[2], [self.batch_size, round(self.height/2), round(self.width/2), 128], self.kernel1, self.kernel1, 2, 2, name='g1_up_deconv3')
            net[3] = batch_norm(net[3], is_train=is_train, name='g1_up_bnorm3')
            net[3] = tf.nn.relu(net[3])
            net[3] = tf.nn.avg_pool(net[3], ksize=[1, 2, 2, 1], strides=[1, 1, 1, 1], padding='SAME', name='g1_up_avg3')
            # 4th de-conv. layer inp : (batch_size, width/2, height/2, 32)
            net[4] = deconv2D(net[3], [self.batch_size, round(self.height), round(self.width), 1], self.kernel1, self.kernel1, 2, 2, name='g1_up_deconv4')
            net[4] = tf.nn.avg_pool(net[4], ksize=[1, 2, 2, 1], strides=[1, 1, 1, 1], padding='SAME', name='g1_up_avg3')
            
            logit = net[4]*1.08
            out = net[4]*1.08
            return out, logit, net
    
    def generator_1_Emb(self, x1, z1, is_train=True, reuse=tf.AUTO_REUSE):
        if not is_train :
         reuse = tf.AUTO_REUSE
        with tf.variable_scope("gen1_Emb", reuse=tf.AUTO_REUSE):
            # generator network(using InfoGAN)
            net_skip1 = {}
            net_skip1[0] = conv2D(x1, 64, self.kernel1, self.kernel1, 1, 1, name='g1_Emb_deconv_x1')
            net_skip1[1] = tf.concat([net_skip1[0], z1*self.alpha], axis=-1)
            # 1st de-conv. layer inp : (self.batch_size, self.height, self.width, 16+1)
            net_skip1[1] = conv2D(net_skip1[1], 64, self.kernel1, self.kernel1, 1, 1, name='g1_Emb_deconv_')
            net_skip1[1] = batch_norm(net_skip1[1], is_train=is_train, name='g1_Emb_bnorm_')
            net_skip1[1] = tf.nn.relu(net_skip1[1])

            net_skip1[2] = conv2D(net_skip1[1], 64, self.kernel1, self.kernel1, 1, 1, name='g1_Emb_deconv1')
            net_skip1[2] = batch_norm(net_skip1[2], is_train=is_train, name='g1_Emb_bnorm1')
            net_skip1[2] = tf.nn.relu(net_skip1[2])
            
            net_skip1[3] = conv2D(net_skip1[2], 64, self.kernel1, self.kernel1, 1, 1, name='g1_Emb_deconv2')
            net_skip1[3] = batch_norm(net_skip1[3], is_train=is_train, name='g1_Emb_bnorm2')
            net_skip1[3] = tf.nn.relu(net_skip1[3])
            
            net_skip1[4] = conv2D(net_skip1[3], 64, self.kernel1, self.kernel1, 1, 1, name='g1_Emb_deconv3')
            net_skip1[4] = batch_norm(net_skip1[4], is_train=is_train, name='g1_Emb_bnorm3')
            net_skip1[4] = tf.nn.relu(net_skip1[4])

            # 2nd de-conv. layer inp : (self.batch_size, self.height, self.width, 4)
            net_skip1[5] = conv2D(net_skip1[4], 1, self.kernel1, self.kernel1, 1, 1, name='g1_Emb_deconv4')
            #net[2] = batch_norm(net[2], is_train=is_train, name='g1_Emb_bnorm4')
            logit = net_skip1[5]
            net_skip1[5] = tf.nn.tanh(logit)
            out = net_skip1[5]
            return out, logit, net_skip1
    
    def generator_2(self, z2, is_train=True, reuse=tf.AUTO_REUSE):
        if not is_train :
         reuse = tf.AUTO_REUSE
        with tf.variable_scope("gen2", reuse=tf.AUTO_REUSE):
            z = z2
            # generator network(using InfoGAN)
            net = {}
            # 1st layer inp : (batch_size, height, width, 1)
            net[1] = conv2D(z, 128, self.kernel2, self.kernel2, 2, 2, name='g2_conv1')
            net[1] = batch_norm(net[1], is_train=is_train, name='g2_bnorm1')
            net[1] = tf.nn.relu(net[1])
            # 2th layer inp : (batch_size, height//2, width//2, 4)
            net[2] = conv2D(net[1], 256, self.kernel2, self.kernel2, 2, 2, name='g2_conv2')
            net[2] = batch_norm(net[2], is_train=is_train, name='g_bnorm2')
            net[2] = tf.nn.relu(net[2])
            # 3th layer inp : (batch_size, height//4, width//4, 16)
            net[3] = conv2D(net[2], 512, self.kernel2, self.kernel2, 2, 2, name='g2_conv3')
            net[3] = batch_norm(net[3], is_train=is_train, name='g2_bnorm3')
            net[3] = tf.nn.relu(net[3])
            # 4th layer inp : (batch_size, height//8, width//8, 64) 
            net[4] = conv2D(net[3], 1, self.kernel2, self.kernel2, 2, 2, name='g2_conv4')
            #net[4] = batch_norm(net[4], is_train=is_train, name='g2_bnorm4')
            #net[4] = tf.nn.relu(net[4])
            # last layer inp : (batch_size, height//16, width//16)
            #net[5] = conv2D(net[4], 1, self.kernel2, self.kernel2, 2, 2, name='g2_conv5')
            logit = net[4]
            net[4] = tf.nn.tanh(logit)
            out = net[4]
            return out, logit, net
    
    def Attack_Identify(self, x1):

        y = x1
        return y

    def Attack_Gaussian_Filtering_test(self, x1, kernel=5, sigma=1):

        y = Gaussian_filtering_test(x1, kernel, sigma, name='Attack_01_k%01d_s%01d'%(kernel, sigma)) 
        return y

    def Attack_Average_Filtering_test(self, x1, kernel=5):

        y = Average_filtering_test(x1, kernel, name='Attack_02_k%01d'%(kernel))
        return y

    def Attack_Salt_and_Pepper(self, x1, p=0.1):

        y = Salt_and_Pepper(x1, p, name='Attack_03_p%01d'%(0.1))
        return y
            
    def Attack_Gaussian_Noise(self, x1, sigma=0.1):

        y = Gaussian_Noise(x1, sigma, name='Attack_04_s%01d'%(0.1))
        return y 

    def Attack_Sharpening(self, x1, center=9):

        y = Sharpening(x1, center, name='Attack_05_s%01d'%(center))
        return y
    
    def Attack_Sharpening_test(self, x1, center=9):

        y = Sharpening_test(x1, center, name='Attack_05_s%01d'%(center))
        return y

    def Attack_Rescaling(self, x1, ratio=0.5):

        y = Rescaling_test(x1, ratio, name='Attack_06_s%01d'%(0.5))
        return y
    
    def Attack_Cropping(self, x1, ratio=0.25):

        y = Cropping(x1, ratio, name='Attack_07_s%01d'%(0.25))
        return y
    
    def Attack_Rotation(self, x1, angle=30):

        y = Rotation_test(x1, name='Attack_08', angle=angle)
        return y
    
    def Attack_Row_and_Column_Removal(self, x1):

        y = Row_and_Column_Removal(x1,  name='Attack_09')
        return y
    
    #def Attack_General_affine_tr(self, x1):
        
        #x1 = tf.multiply(tf.divide(tf.add(x1, 1), 2), 255)
        #x1 = tf.clip_by_value(x1, 0, 255)
        #y = General_affine_tr(x1, sigma, name='Attack_10')
        #y = tf.clip_by_value(tf.subtract(tf.divide(y, 127.5), 1), -1, 1) 
        #return y
    
    def Attack_Contrast(self, x1, factor=5):

        y = Contrast(x1, factor, name='Attack_11_s%01d'%(factor))
        return y
    
    def Attack_Gamma_Correction(self, x1, gamma=0.3):

        y = Gamma_Correction(x1, gamma, name='Attack_12_s%01d'%(gamma))
        return y
    
    def Attack_JPEG(self, x1, quality=50):

        y = JPEG(x1, quality, name='Attack_13_s%01d'%(50))
        return y
    
    def Attack_Grid_Crop(self, x1, ratio=0.5):

        y = Grid_Crop_test(x1, ratio=ratio, name='Attack_14_')
        return y

    def Attack_Dropout(self, x1, x2, p=0.3):

        y = Dropout_test(x1, x2, p, name='Attack_15')
        return y
    
    def Attack_Median_Filtering_test(self, x1, kernel):

        y = Median_filter_test(x1, kernel, name='Attack_16_k%01d'%(kernel))
        return y


    def NCC(self, z, x):
        # BER
        z_1_ = tf.add(z, 1)
        z_1_ = tf.divide(z_1_, 2)
        z_1_ = tf.clip_by_value(z_1_, 0, 1)
        z_1_ = tf.round(z_1_)

        G_test_2_ = tf.add(x, 1)
        G_test_2_ = tf.divide(G_test_2_, 2)
        G_test_2_ = tf.clip_by_value(G_test_2_, 0, 1)
        G_test_2_ = tf.round(G_test_2_)

        BER = tf.equal(z_1_, G_test_2_)
        BER = tf.cast(BER, tf.float32)
        BER = tf.reduce_mean(BER)

        return 100-100*BER

    def build_model(self):
        self.x_1 = tf.placeholder(tf.float32, [self.batch_size, self.height, self.width, 1], name='Host_ori')
        self.z_1 = tf.placeholder(tf.float32, [self.batch_size, round(self.height/16), round(self.width/16), 1], name='WM')

        def sigmoid_cross_entropy_with_logits(x, y):
              try:
                return tf.nn.sigmoid_cross_entropy_with_logits(logits=x, labels=y)
              except:
                return tf.nn.sigmoid_cross_entropy_with_logits(logits=x, targets=y)

        # Training graph ################################################################################
        # output of D for fake data
        # self.concated = tf.concat([self.x_1,self.x_1],axis=-1)
        G_up_fake_1, _, net_up_1 =self.generator_1_up(self.z_1, is_train=True, reuse=tf.AUTO_REUSE)
        self.G_fake_1, _, net_1 =self.generator_1_Emb(self.x_1, G_up_fake_1, is_train=True, reuse=tf.AUTO_REUSE)
        #print(self.G_fake_1)

        # Attack 
        self.G_fake_1_ = self.G_fake_1

        # output of D for fake data
        self.G_fake_2_, _, _ = self.generator_2(self.G_fake_1_, is_train=True, reuse=tf.AUTO_REUSE)
        #self.G_fake_2_test, _, _ = self.generator_2(self.G_fake_1_test, is_train=False, reuse=True)

        # loss for G
        self.invisibility = tf.reduce_mean(tf.multiply(tf.subtract(self.G_fake_1, self.x_1), tf.subtract(self.G_fake_1, self.x_1)))
        self.robustness_ = tf.reduce_mean(tf.abs(tf.subtract(self.G_fake_2_, self.z_1)))

        # loss for Embedder
        self.G1_loss_total_ = (self.lambda1*self.invisibility) + (self.lambda2*self.robustness_) 

        # loss for Extractor
        self.G2_loss_total_ = self.lambda3*self.robustness_

        # optimizers
        t_vars = tf.trainable_variables()
        self.G_vars_1 = [var for var in t_vars if 'gen1' in var.name]
        self.G_vars_2 = [var for var in t_vars if 'gen2' in var.name]

        with tf.control_dependencies(tf.get_collection(tf.GraphKeys.UPDATE_OPS)):
            self.G_optim_1_ = tf.train.AdamOptimizer(self.g1_lr, beta1=self.beta1, beta2=self.beta2) \
                            .minimize(self.G1_loss_total_, var_list=self.G_vars_1)
            self.G_optim_2_ = tf.train.AdamOptimizer(self.gan2_lr, beta1=self.beta1, beta2=self.beta2) \
                            .minimize(self.G2_loss_total_, var_list=self.G_vars_2)
    
    def build_model_test(self):
        self.WMed = tf.placeholder(tf.float32, [self.batch_size, self.height, self.width, 1], name='WMed')
        self.x_2 = tf.placeholder(tf.complex64, [self.batch_size, self.height, self.width, 1], name='Host_imag')
        # Test graph ####################################################################################
        G_up_test_1, _, self.G_up_features_1 = self.generator_1_up(self.z_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_1, _, self.G_features_1 = self.generator_1_Emb(self.x_1, G_up_test_1, is_train=False, reuse=tf.AUTO_REUSE)

        #self.G_fake_1_test = self.WMe
        
        # PSNR
        # a = tf.sqrt(tf.add(x1,x2))
        # a = tf.multiply(a,a)
        # b = tf.norm(tf.subtract(self.x_1,self.G_test_1))
        # b = tf.multiply(b,b)
        # numerator = tf.log(tf.divide(a,b))
        # denominator = tf.log(tf.constant(10, dtype=numerator.dtype))
        # self.csnr_test = tf.multiply(tf.constant(10, dtype=numerator.dtype),tf.divide(numerator,denominator))
        # self.psnr_test = tf.reduce_mean(tf.image.psnr(self.G_test_1, self.x_1, max_val=2))
        # x1 = tf.add(self.x_1,tf.constant(1,dtype=tf.float32))
        # x1 = tf.divide(x1,tf.constant(2,dtype=tf.float32))
        # x1 = tf.multiply(x1,self.in_maxmin)
        # x1 = tf.add(x1,self.in_min)
        #
        # x2 = tf.add(self.G_test_1, tf.constant(1, dtype=tf.float32))
        # x2 = tf.divide(x2, tf.constant(2, dtype=tf.float32))
        # x2 = tf.multiply(x2, self.in_maxmin)
        # x2 = tf.add(x2, self.in_min)

        x1 = tf.complex(self.x_1, tf.constant(0.0))
        x2 = tf.complex(self.G_test_1, tf.constant(0.0))
        com1_ = tf.exp(tf.multiply(tf.constant(1j, dtype=tf.complex64), self.x_2))
        com1 = tf.multiply(com1_, x1)
        com2 = tf.multiply(com1_, x2)
        # PSNR
        # a = tf.sqrt(tf.add(x1,x2))
        a = tf.norm(com1)
        a = tf.multiply(a,a)
        b = tf.norm(tf.subtract(com1,com2))
        b = tf.multiply(b,b)
        numerator = tf.log(tf.divide(a,b))
        denominator = tf.log(tf.constant(10, dtype=numerator.dtype))
        self.csnr_test = tf.multiply(tf.constant(10, dtype=numerator.dtype),tf.divide(numerator,denominator))
        self.psnr_test = tf.reduce_mean(tf.image.psnr(self.G_test_1, self.x_1, max_val=2))

        # Attack 
        self.G_test_1_ = self.Attack_Identify(self.G_test_1)
        self.G_test_1_A1_1 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=3, sigma=0.5)
        self.G_test_1_A1_2 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=5, sigma=1)
        self.G_test_1_A1_3 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=7, sigma=1.5)
        self.G_test_1_A1_4 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=9, sigma=2)
        self.G_test_1_A1_5 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=11, sigma=2.5)
        self.G_test_1_A1_6 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=13, sigma=3)
        self.G_test_1_A1_7 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=15, sigma=3.5)
        self.G_test_1_A1_8 = self.Attack_Gaussian_Filtering_test(self.G_test_1, kernel=17, sigma=4)
        self.G_test_1_A2_1 = self.Attack_Average_Filtering_test(self.G_test_1, kernel=5)
        self.G_test_1_A2_2 = self.Attack_Average_Filtering_test(self.G_test_1, kernel=3)
        self.G_test_1_A3_1 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.1)
        self.G_test_1_A3_2 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.09)
        self.G_test_1_A3_3 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.08)
        self.G_test_1_A3_4 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.07)
        self.G_test_1_A3_5 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.06)
        self.G_test_1_A3_6 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.05)
        self.G_test_1_A3_7 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.04)
        self.G_test_1_A3_8 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.03)
        self.G_test_1_A3_9 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.02)
        self.G_test_1_A3_10 = self.Attack_Salt_and_Pepper(self.G_test_1, 0.01)
        self.G_test_1_A4_1 = self.Attack_Gaussian_Noise(self.G_test_1, 0.1)
        self.G_test_1_A4_2 = self.Attack_Gaussian_Noise(self.G_test_1, 0.08)
        self.G_test_1_A4_3 = self.Attack_Gaussian_Noise(self.G_test_1, 0.05)
        self.G_test_1_A4_4 = self.Attack_Gaussian_Noise(self.G_test_1, 0.03)
        self.G_test_1_A4_5 = self.Attack_Gaussian_Noise(self.G_test_1, 0.01)
        self.G_test_1_A5_1 = self.Attack_Sharpening_test(self.G_test_1, 9)
        self.G_test_1_A5_2 = self.Attack_Sharpening_test(self.G_test_1, 5)
        self.G_test_1_A6_1 = self.Attack_Rescaling(self.G_test_1, 0.4)
        self.G_test_1_A6_2 = self.Attack_Rescaling(self.G_test_1, 0.5)
        self.G_test_1_A6_3 = self.Attack_Rescaling(self.G_test_1, 0.6)
        self.G_test_1_A6_4 = self.Attack_Rescaling(self.G_test_1, 0.7)
        self.G_test_1_A6_5 = self.Attack_Rescaling(self.G_test_1, 0.8)
        self.G_test_1_A6_6 = self.Attack_Rescaling(self.G_test_1, 0.9)
        self.G_test_1_A7_1 = self.Attack_Cropping(self.G_test_1, 0.9)
        self.G_test_1_A7_2 = self.Attack_Cropping(self.G_test_1, 0.8)
        self.G_test_1_A7_3 = self.Attack_Cropping(self.G_test_1, 0.7)
        self.G_test_1_A7_4 = self.Attack_Cropping(self.G_test_1, 0.6)
        self.G_test_1_A7_5 = self.Attack_Cropping(self.G_test_1, 0.5)
        self.G_test_1_A7_6 = self.Attack_Cropping(self.G_test_1, 0.4)
        self.G_test_1_A7_7 = self.Attack_Cropping(self.G_test_1, 0.3)
        self.G_test_1_A7_8 = self.Attack_Cropping(self.G_test_1, 0.2)
        self.G_test_1_A7_9 = self.Attack_Cropping(self.G_test_1, 0.1)
        self.G_test_1_A8_1 = self.Attack_Rotation(self.G_test_1, 15)
        self.G_test_1_A8_2 = self.Attack_Rotation(self.G_test_1, 30)
        self.G_test_1_A8_3 = self.Attack_Rotation(self.G_test_1, 45)
        self.G_test_1_A8_4 = self.Attack_Rotation(self.G_test_1, 60)
        self.G_test_1_A8_5 = self.Attack_Rotation(self.G_test_1, 75)
        self.G_test_1_A8_6 = self.Attack_Rotation(self.G_test_1, 90)
        #self.G_test_1_A9 = self.Attack_Row_and_Column_Removal(self.G_test_1)
        #self.G_test_1_A11 = self.Attack_Contrast(self.G_test_1)
        self.G_test_1_A13_1 = self.Attack_JPEG(self.G_test_1, 10)
        self.G_test_1_A13_2 = self.Attack_JPEG(self.G_test_1, 20)
        self.G_test_1_A13_3 = self.Attack_JPEG(self.G_test_1, 30)
        self.G_test_1_A13_4 = self.Attack_JPEG(self.G_test_1, 40)
        self.G_test_1_A13_5 = self.Attack_JPEG(self.G_test_1, 50)
        self.G_test_1_A13_6 = self.Attack_JPEG(self.G_test_1, 60)
        self.G_test_1_A13_7 = self.Attack_JPEG(self.G_test_1, 70)
        self.G_test_1_A13_8 = self.Attack_JPEG(self.G_test_1, 80)
        self.G_test_1_A13_9 = self.Attack_JPEG(self.G_test_1, 90)
        self.G_test_1_A13_10 = self.Attack_JPEG(self.G_test_1, 100)
        self.G_test_1_A14_1 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.035)
        self.G_test_1_A14_2 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.1)
        self.G_test_1_A14_3 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.2)
        self.G_test_1_A14_4 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.3)
        self.G_test_1_A14_5 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.4)
        self.G_test_1_A14_6 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.5)
        self.G_test_1_A14_7 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.6)
        self.G_test_1_A14_8 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.7)
        self.G_test_1_A14_9 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.8)
        self.G_test_1_A14_10 = self.Attack_Grid_Crop(self.G_test_1, ratio=0.9)
        self.G_test_1_A15_1 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.1)
        self.G_test_1_A15_2 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.2)
        self.G_test_1_A15_3 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.3)
        self.G_test_1_A15_4 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.4)
        self.G_test_1_A15_5 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.5)
        self.G_test_1_A15_6 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.6)
        self.G_test_1_A15_7 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.7)
        self.G_test_1_A15_8 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.8)
        self.G_test_1_A15_9 = self.Attack_Dropout(self.G_test_1, self.x_1, 0.9)
        self.G_test_1_A16_1 = self.Attack_Median_Filtering_test(self.G_test_1, 3)
        self.G_test_1_A16_2 = self.Attack_Median_Filtering_test(self.G_test_1, 5)
        self.G_test_1_A16_3 = self.Attack_Median_Filtering_test(self.G_test_1, 7)

        # output of D for fake data
        self.G_test_2_, _, _ = self.generator_2(self.G_test_1_, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_1, _, _ = self.generator_2(self.G_test_1_A1_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_2, _, _ = self.generator_2(self.G_test_1_A1_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_3, _, _ = self.generator_2(self.G_test_1_A1_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_4, _, _ = self.generator_2(self.G_test_1_A1_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_5, _, _ = self.generator_2(self.G_test_1_A1_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_6, _, _ = self.generator_2(self.G_test_1_A1_6, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_7, _, _ = self.generator_2(self.G_test_1_A1_7, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A1_8, _, _ = self.generator_2(self.G_test_1_A1_8, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A2_1, _, _ = self.generator_2(self.G_test_1_A2_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A2_2, _, _ = self.generator_2(self.G_test_1_A2_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_1, _, _ = self.generator_2(self.G_test_1_A3_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_2, _, _ = self.generator_2(self.G_test_1_A3_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_3, _, _ = self.generator_2(self.G_test_1_A3_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_4, _, _ = self.generator_2(self.G_test_1_A3_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_5, _, _ = self.generator_2(self.G_test_1_A3_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_6, _, _ = self.generator_2(self.G_test_1_A3_6, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_7, _, _ = self.generator_2(self.G_test_1_A3_7, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_8, _, _ = self.generator_2(self.G_test_1_A3_8, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_9, _, _ = self.generator_2(self.G_test_1_A3_9, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A3_10, _, _ = self.generator_2(self.G_test_1_A3_10, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A4_1, _, _ = self.generator_2(self.G_test_1_A4_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A4_2, _, _ = self.generator_2(self.G_test_1_A4_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A4_3, _, _ = self.generator_2(self.G_test_1_A4_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A4_4, _, _ = self.generator_2(self.G_test_1_A4_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A4_5, _, _ = self.generator_2(self.G_test_1_A4_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A5_1, _, _ = self.generator_2(self.G_test_1_A5_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A5_2, _, _ = self.generator_2(self.G_test_1_A5_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A6_1, _, _ = self.generator_2(self.G_test_1_A6_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A6_2, _, _ = self.generator_2(self.G_test_1_A6_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A6_3, _, _ = self.generator_2(self.G_test_1_A6_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A6_4, _, _ = self.generator_2(self.G_test_1_A6_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A6_5, _, _ = self.generator_2(self.G_test_1_A6_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A6_6, _, _ = self.generator_2(self.G_test_1_A6_6, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_1, _, _ = self.generator_2(self.G_test_1_A7_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_2, _, _ = self.generator_2(self.G_test_1_A7_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_3, _, _ = self.generator_2(self.G_test_1_A7_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_4, _, _ = self.generator_2(self.G_test_1_A7_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_5, _, _ = self.generator_2(self.G_test_1_A7_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_6, _, _ = self.generator_2(self.G_test_1_A7_6, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_7, _, _ = self.generator_2(self.G_test_1_A7_7, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_8, _, _ = self.generator_2(self.G_test_1_A7_8, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A7_9, _, _ = self.generator_2(self.G_test_1_A7_9, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A8_1, _, _ = self.generator_2(self.G_test_1_A8_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A8_2, _, _ = self.generator_2(self.G_test_1_A8_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A8_3, _, _ = self.generator_2(self.G_test_1_A8_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A8_4, _, _ = self.generator_2(self.G_test_1_A8_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A8_5, _, _ = self.generator_2(self.G_test_1_A8_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A8_6, _, _ = self.generator_2(self.G_test_1_A8_6, is_train=False, reuse=tf.AUTO_REUSE)
        #self.G_test_2_A9, _, _ = self.generator_2(self.G_test_1_A9, is_train=False, reuse=True)
        #self.G_test_2_A11, _, _ = self.generator_2(self.G_test_1_A11, is_train=False, reuse=True)
        self.G_test_2_A13_1, _, _ = self.generator_2(self.G_test_1_A13_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_2, _, _ = self.generator_2(self.G_test_1_A13_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_3, _, _ = self.generator_2(self.G_test_1_A13_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_4, _, _ = self.generator_2(self.G_test_1_A13_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_5, _, _ = self.generator_2(self.G_test_1_A13_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_6, _, _ = self.generator_2(self.G_test_1_A13_6, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_7, _, _ = self.generator_2(self.G_test_1_A13_7, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_8, _, _ = self.generator_2(self.G_test_1_A13_8, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_9, _, _ = self.generator_2(self.G_test_1_A13_9, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A13_10, _, _ = self.generator_2(self.G_test_1_A13_10, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_1, _, _ = self.generator_2(self.G_test_1_A14_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_2, _, _ = self.generator_2(self.G_test_1_A14_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_3, _, _ = self.generator_2(self.G_test_1_A14_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_4, _, _ = self.generator_2(self.G_test_1_A14_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_5, _, _ = self.generator_2(self.G_test_1_A14_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_6, _, _ = self.generator_2(self.G_test_1_A14_6, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_7, _, _ = self.generator_2(self.G_test_1_A14_7, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_8, _, _ = self.generator_2(self.G_test_1_A14_8, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_9, _, _ = self.generator_2(self.G_test_1_A14_9, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A14_10, _, _ = self.generator_2(self.G_test_1_A14_10, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_1, _, _ = self.generator_2(self.G_test_1_A15_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_2, _, _ = self.generator_2(self.G_test_1_A15_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_3, _, _ = self.generator_2(self.G_test_1_A15_3, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_4, _, _ = self.generator_2(self.G_test_1_A15_4, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_5, _, _ = self.generator_2(self.G_test_1_A15_5, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_6, _, _ = self.generator_2(self.G_test_1_A15_6, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_7, _, _ = self.generator_2(self.G_test_1_A15_7, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_8, _, _ = self.generator_2(self.G_test_1_A15_8, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A15_9, _, _ = self.generator_2(self.G_test_1_A15_9, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A16_1, _, _ = self.generator_2(self.G_test_1_A16_1, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A16_2, _, _ = self.generator_2(self.G_test_1_A16_2, is_train=False, reuse=tf.AUTO_REUSE)
        self.G_test_2_A16_3, _, _ = self.generator_2(self.G_test_1_A16_3, is_train=False, reuse=tf.AUTO_REUSE)

        # NCC
        self.NC_WM_t_ = self.NCC(self.z_1, self.G_test_2_)
        self.NC_WM_t_A1_1 = self.NCC(self.z_1, self.G_test_2_A1_1)
        self.NC_WM_t_A1_2 = self.NCC(self.z_1, self.G_test_2_A1_2)
        self.NC_WM_t_A1_3 = self.NCC(self.z_1, self.G_test_2_A1_3)
        self.NC_WM_t_A1_4 = self.NCC(self.z_1, self.G_test_2_A1_4)
        self.NC_WM_t_A1_5 = self.NCC(self.z_1, self.G_test_2_A1_5)
        self.NC_WM_t_A1_6 = self.NCC(self.z_1, self.G_test_2_A1_6)
        self.NC_WM_t_A1_7 = self.NCC(self.z_1, self.G_test_2_A1_7)
        self.NC_WM_t_A1_8 = self.NCC(self.z_1, self.G_test_2_A1_8)
        self.NC_WM_t_A2_1 = self.NCC(self.z_1, self.G_test_2_A2_1)
        self.NC_WM_t_A2_2 = self.NCC(self.z_1, self.G_test_2_A2_2)
        self.NC_WM_t_A3_1 = self.NCC(self.z_1, self.G_test_2_A3_1)
        self.NC_WM_t_A3_2 = self.NCC(self.z_1, self.G_test_2_A3_2)
        self.NC_WM_t_A3_3 = self.NCC(self.z_1, self.G_test_2_A3_3)
        self.NC_WM_t_A3_4 = self.NCC(self.z_1, self.G_test_2_A3_4)
        self.NC_WM_t_A3_5 = self.NCC(self.z_1, self.G_test_2_A3_5)
        self.NC_WM_t_A3_6 = self.NCC(self.z_1, self.G_test_2_A3_6)
        self.NC_WM_t_A3_7 = self.NCC(self.z_1, self.G_test_2_A3_7)
        self.NC_WM_t_A3_8 = self.NCC(self.z_1, self.G_test_2_A3_8)
        self.NC_WM_t_A3_9 = self.NCC(self.z_1, self.G_test_2_A3_9)
        self.NC_WM_t_A3_10 = self.NCC(self.z_1, self.G_test_2_A3_10)
        self.NC_WM_t_A4_1 = self.NCC(self.z_1, self.G_test_2_A4_1)
        self.NC_WM_t_A4_2 = self.NCC(self.z_1, self.G_test_2_A4_2)
        self.NC_WM_t_A4_3 = self.NCC(self.z_1, self.G_test_2_A4_3)
        self.NC_WM_t_A4_4 = self.NCC(self.z_1, self.G_test_2_A4_4)
        self.NC_WM_t_A4_5 = self.NCC(self.z_1, self.G_test_2_A4_5)
        self.NC_WM_t_A5_1 = self.NCC(self.z_1, self.G_test_2_A5_1)
        self.NC_WM_t_A5_2 = self.NCC(self.z_1, self.G_test_2_A5_2)
        self.NC_WM_t_A6_1 = self.NCC(self.z_1, self.G_test_2_A6_1)
        self.NC_WM_t_A6_2 = self.NCC(self.z_1, self.G_test_2_A6_2)
        self.NC_WM_t_A6_3 = self.NCC(self.z_1, self.G_test_2_A6_3)
        self.NC_WM_t_A6_4 = self.NCC(self.z_1, self.G_test_2_A6_4)
        self.NC_WM_t_A6_5 = self.NCC(self.z_1, self.G_test_2_A6_5)
        self.NC_WM_t_A6_6 = self.NCC(self.z_1, self.G_test_2_A6_6)
        self.NC_WM_t_A7_1 = self.NCC(self.z_1, self.G_test_2_A7_1)
        self.NC_WM_t_A7_2 = self.NCC(self.z_1, self.G_test_2_A7_2)
        self.NC_WM_t_A7_3 = self.NCC(self.z_1, self.G_test_2_A7_3)
        self.NC_WM_t_A7_4 = self.NCC(self.z_1, self.G_test_2_A7_4)
        self.NC_WM_t_A7_5 = self.NCC(self.z_1, self.G_test_2_A7_5)
        self.NC_WM_t_A7_6 = self.NCC(self.z_1, self.G_test_2_A7_6)
        self.NC_WM_t_A7_7 = self.NCC(self.z_1, self.G_test_2_A7_7)
        self.NC_WM_t_A7_8 = self.NCC(self.z_1, self.G_test_2_A7_8)
        self.NC_WM_t_A7_9 = self.NCC(self.z_1, self.G_test_2_A7_9)
        self.NC_WM_t_A8_1 = self.NCC(self.z_1, self.G_test_2_A8_1)
        self.NC_WM_t_A8_2 = self.NCC(self.z_1, self.G_test_2_A8_2)
        self.NC_WM_t_A8_3 = self.NCC(self.z_1, self.G_test_2_A8_3)
        self.NC_WM_t_A8_4 = self.NCC(self.z_1, self.G_test_2_A8_4)
        self.NC_WM_t_A8_5 = self.NCC(self.z_1, self.G_test_2_A8_5)
        self.NC_WM_t_A8_6 = self.NCC(self.z_1, self.G_test_2_A8_6)
        #self.NC_WM_t_A9 = self.NCC(self.z_1, self.G_test_2_A9)
        #self.NC_WM_t_A11 = self.NCC(self.z_1, self.G_test_2_A11)
        self.NC_WM_t_A13_1 = self.NCC(self.z_1, self.G_test_2_A13_1)
        self.NC_WM_t_A13_2 = self.NCC(self.z_1, self.G_test_2_A13_2)
        self.NC_WM_t_A13_3 = self.NCC(self.z_1, self.G_test_2_A13_3)
        self.NC_WM_t_A13_4 = self.NCC(self.z_1, self.G_test_2_A13_4)
        self.NC_WM_t_A13_5 = self.NCC(self.z_1, self.G_test_2_A13_5)
        self.NC_WM_t_A13_6 = self.NCC(self.z_1, self.G_test_2_A13_6)
        self.NC_WM_t_A13_7 = self.NCC(self.z_1, self.G_test_2_A13_7)
        self.NC_WM_t_A13_8 = self.NCC(self.z_1, self.G_test_2_A13_8)
        self.NC_WM_t_A13_9 = self.NCC(self.z_1, self.G_test_2_A13_9)
        self.NC_WM_t_A13_10 = self.NCC(self.z_1, self.G_test_2_A13_10)
        self.NC_WM_t_A14_1 = self.NCC(self.z_1, self.G_test_2_A14_1)
        self.NC_WM_t_A14_2 = self.NCC(self.z_1, self.G_test_2_A14_2)
        self.NC_WM_t_A14_3 = self.NCC(self.z_1, self.G_test_2_A14_3)
        self.NC_WM_t_A14_4 = self.NCC(self.z_1, self.G_test_2_A14_4)
        self.NC_WM_t_A14_5 = self.NCC(self.z_1, self.G_test_2_A14_5)
        self.NC_WM_t_A14_6 = self.NCC(self.z_1, self.G_test_2_A14_6)
        self.NC_WM_t_A14_7 = self.NCC(self.z_1, self.G_test_2_A14_7)
        self.NC_WM_t_A14_8 = self.NCC(self.z_1, self.G_test_2_A14_8)
        self.NC_WM_t_A14_9 = self.NCC(self.z_1, self.G_test_2_A14_9)
        self.NC_WM_t_A14_10 = self.NCC(self.z_1, self.G_test_2_A14_10)
        self.NC_WM_t_A15_1 = self.NCC(self.z_1, self.G_test_2_A15_1)
        self.NC_WM_t_A15_2 = self.NCC(self.z_1, self.G_test_2_A15_2)
        self.NC_WM_t_A15_3 = self.NCC(self.z_1, self.G_test_2_A15_3)
        self.NC_WM_t_A15_4 = self.NCC(self.z_1, self.G_test_2_A15_4)
        self.NC_WM_t_A15_5 = self.NCC(self.z_1, self.G_test_2_A15_5)
        self.NC_WM_t_A15_6 = self.NCC(self.z_1, self.G_test_2_A15_6)
        self.NC_WM_t_A15_7 = self.NCC(self.z_1, self.G_test_2_A15_7)
        self.NC_WM_t_A15_8 = self.NCC(self.z_1, self.G_test_2_A15_8)
        self.NC_WM_t_A15_9 = self.NCC(self.z_1, self.G_test_2_A15_9)
        self.NC_WM_t_A16_1 = self.NCC(self.z_1, self.G_test_2_A16_1)
        self.NC_WM_t_A16_2 = self.NCC(self.z_1, self.G_test_2_A16_2)
        self.NC_WM_t_A16_3 = self.NCC(self.z_1, self.G_test_2_A16_3)



    def test(self, dataset_name, ep, listdir): #, data

        # dataset load
        # Host_test = np.load('../01_Hologram_h5/data2/'+data)
        # Host_test = Host_test.reshape([-1, 1, self.height, self.width, 1])

        # data = h5py.File('../01_Hologram_h5/data2/01_ERC_hologram_test', 'r')
        # Host_test = list(data['data'])
        # Host_test = np.asarray(Host_test)#[:self.batch_size][:][:]
        # Host_test = Host_test.reshape([-1, self.height, self.width, 1])
        data = h5py.File('./20210301_global_std_test_amp', 'r')
        Host_test = np.asarray(data['data']).reshape([-1, self.height, self.width, 1])

        data0 = h5py.File('./20210301_global_std_test_phase', 'r')
        Host_imag = np.asarray(data0['data']).reshape([-1, self.height, self.width, 1])

        print(Host_test.shape)

        

        # session creation
        run_config = tf.ConfigProto()
        run_config.gpu_options.allow_growth = True
        self.sess = tf.Session(config=run_config)

        self.sess.run(tf.global_variables_initializer())

        # saver to save model
        self.saver = tf.train.Saver()

        # check-point restore if it exists
        load_flag, load_count = self.load_checkpoint('./ckt', dataset_name, ep)
        if (load_flag):
            start_epoch = load_count//50
            #start_batch_id = load_count - start_epoch * num_batches
            counter = load_count
        else:
            start_epoch = 0
            #start_batch_id = 0
            counter = 0

        start_time = time.time() 

        np.random.seed(1)
        WM_data_a = np.random.uniform(0, 1, [1, round(self.height/16), round(self.width/16), 1])
        WM_data_a = np.round(WM_data_a)
        WM_data_b = WM_data_a-1
        WM_data = WM_data_a + WM_data_b
        WM_data = np.repeat(WM_data, self.batch_size, 0)

        
        print(dataset_name)
        # print(data)
        dataset_name_ = dataset_name+'/'+'128x128_s1/'+ '%s/' %(listdir)

        # directory check
        directory = './result/'+dataset_name_
        if not os.path.exists(directory):
            os.makedirs(directory)

        #NC_WM_t_A5_2 \ self.NC_WM_t_A5_2
        # test generator network

        Final_Result = np.zeros([round(Host_test.shape[0]/self.batch_size), 83])

        For_Recon = []
        for i in range(round(Host_test.shape[0]/self.batch_size)) :

            WMed_, g_test_2, CSNR_WMed_t_, PSNR_WMed_t_,\
            NC_WM_t_, NC_WM_t_A1_1, NC_WM_t_A1_2, NC_WM_t_A1_3, NC_WM_t_A1_4, \
            NC_WM_t_A1_5, NC_WM_t_A1_6, NC_WM_t_A1_7, NC_WM_t_A1_8, NC_WM_t_A2_1, NC_WM_t_A2_2, \
            NC_WM_t_A3_1, NC_WM_t_A3_2, NC_WM_t_A3_3, NC_WM_t_A3_4, NC_WM_t_A3_5, \
            NC_WM_t_A3_6, NC_WM_t_A3_7, NC_WM_t_A3_8, NC_WM_t_A3_9, NC_WM_t_A3_10, \
            NC_WM_t_A4_1, NC_WM_t_A4_2, NC_WM_t_A4_3, NC_WM_t_A4_4, NC_WM_t_A4_5, \
            NC_WM_t_A5_1, NC_WM_t_A5_2 \
                = self.sess.run([self.G_test_1, self.G_test_2_, self.csnr_test, self.psnr_test, \
                                 self.NC_WM_t_, self.NC_WM_t_A1_1, self.NC_WM_t_A1_2, self.NC_WM_t_A1_3, self.NC_WM_t_A1_4, \
                                 self.NC_WM_t_A1_5, self.NC_WM_t_A1_6, self.NC_WM_t_A1_7, self.NC_WM_t_A1_8, self.NC_WM_t_A2_1, self.NC_WM_t_A2_2, \
                                 self.NC_WM_t_A3_1, self.NC_WM_t_A3_2, self.NC_WM_t_A3_3, self.NC_WM_t_A3_4, self.NC_WM_t_A3_5, \
                                 self.NC_WM_t_A3_6, self.NC_WM_t_A3_7, self.NC_WM_t_A3_8, self.NC_WM_t_A3_9, self.NC_WM_t_A3_10, \
                                 self.NC_WM_t_A4_1, self.NC_WM_t_A4_2, self.NC_WM_t_A4_3, self.NC_WM_t_A4_4, self.NC_WM_t_A4_5, \
                                 self.NC_WM_t_A5_1, self.NC_WM_t_A5_2], 
                feed_dict={
                self.x_1 : Host_test[i*self.batch_size:(i+1)*self.batch_size].astype(np.float32),
                self.x_2 : Host_imag[i*self.batch_size:(i+1)*self.batch_size].astype(np.float32),
                self.z_1 : WM_data.astype(np.float32)
            })
    
    
            # test generator network
            NC_WM_t_A6_1, NC_WM_t_A6_2, NC_WM_t_A6_3, NC_WM_t_A6_4, NC_WM_t_A6_5, NC_WM_t_A6_6, \
            NC_WM_t_A7_1, NC_WM_t_A7_2, NC_WM_t_A7_3, NC_WM_t_A7_4, \
            NC_WM_t_A7_5, NC_WM_t_A7_6, NC_WM_t_A7_7, NC_WM_t_A7_8, NC_WM_t_A7_9, \
            NC_WM_t_A8_1, NC_WM_t_A8_2, NC_WM_t_A8_3, NC_WM_t_A8_4, NC_WM_t_A8_5, NC_WM_t_A8_6, \
            NC_WM_t_A13_1, NC_WM_t_A13_2, NC_WM_t_A13_3, NC_WM_t_A13_4, NC_WM_t_A13_5, NC_WM_t_A13_6, \
            NC_WM_t_A13_7, NC_WM_t_A13_8, NC_WM_t_A13_9, NC_WM_t_A13_10, \
            NC_WM_t_A14_1, NC_WM_t_A14_2, NC_WM_t_A14_3, NC_WM_t_A14_4, NC_WM_t_A14_5, NC_WM_t_A14_6, \
            NC_WM_t_A14_7, NC_WM_t_A14_8, NC_WM_t_A14_9, NC_WM_t_A14_10, \
            NC_WM_t_A15_1, NC_WM_t_A15_2, NC_WM_t_A15_3, NC_WM_t_A15_4, NC_WM_t_A15_5, \
            NC_WM_t_A15_6, NC_WM_t_A15_7, NC_WM_t_A15_8, NC_WM_t_A15_9, \
            NC_WM_t_A16_1, NC_WM_t_A16_2, NC_WM_t_A16_3 \
                = self.sess.run([self.NC_WM_t_A6_1, self.NC_WM_t_A6_2, self.NC_WM_t_A6_3, self.NC_WM_t_A6_4, self.NC_WM_t_A6_5, self.NC_WM_t_A6_6,\
                                 self.NC_WM_t_A7_1, self.NC_WM_t_A7_2, self.NC_WM_t_A7_3, self.NC_WM_t_A7_4, \
                                 self.NC_WM_t_A7_5, self.NC_WM_t_A7_6, self.NC_WM_t_A7_7, self.NC_WM_t_A7_8, self.NC_WM_t_A7_9, \
                                 self.NC_WM_t_A8_1, self.NC_WM_t_A8_2, self.NC_WM_t_A8_3, self.NC_WM_t_A8_4, self.NC_WM_t_A8_5, self.NC_WM_t_A8_6, \
                                 self.NC_WM_t_A13_1, self.NC_WM_t_A13_2, self.NC_WM_t_A13_3, self.NC_WM_t_A13_4, self.NC_WM_t_A13_5, self.NC_WM_t_A13_6, \
                                 self.NC_WM_t_A13_7, self.NC_WM_t_A13_8, self.NC_WM_t_A13_9, self.NC_WM_t_A13_10, \
                                 self.NC_WM_t_A14_1, self.NC_WM_t_A14_2, self.NC_WM_t_A14_3, self.NC_WM_t_A14_4, self.NC_WM_t_A14_5, self.NC_WM_t_A14_6,\
                                 self.NC_WM_t_A14_7, self.NC_WM_t_A14_8, self.NC_WM_t_A14_9, self.NC_WM_t_A14_10, \
                                 self.NC_WM_t_A15_1, self.NC_WM_t_A15_2, self.NC_WM_t_A15_3, self.NC_WM_t_A15_4, self.NC_WM_t_A15_5, \
                                 self.NC_WM_t_A15_6, self.NC_WM_t_A15_7, self.NC_WM_t_A15_8, self.NC_WM_t_A15_9, \
                                 self.NC_WM_t_A16_1, self.NC_WM_t_A16_2, self.NC_WM_t_A16_3], 
                feed_dict={
                self.x_1 : Host_test[i*self.batch_size:(i+1)*self.batch_size].astype(np.float32),
                self.z_1 : WM_data.astype(np.float32)
            })
    
            G_test_1_,     G_test_1_A1_1, G_test_1_A1_2, G_test_1_A1_3, G_test_1_A1_4, \
            G_test_1_A1_5, G_test_1_A1_6, G_test_1_A1_7, G_test_1_A1_8, G_test_1_A2_1, G_test_1_A2_2, \
            G_test_1_A3_1, G_test_1_A3_2, G_test_1_A3_3, G_test_1_A3_4, G_test_1_A3_5, \
            G_test_1_A3_6, G_test_1_A3_7, G_test_1_A3_8, G_test_1_A3_9, G_test_1_A3_10, \
            G_test_1_A4_1, G_test_1_A4_2, G_test_1_A4_3, G_test_1_A4_4, G_test_1_A4_5, \
            G_test_1_A5_1, G_test_1_A5_2 \
            = self.sess.run([self.G_test_1_,     self.G_test_1_A1_1, self.G_test_1_A1_2, self.G_test_1_A1_3, self.G_test_1_A1_4, \
                             self.G_test_1_A1_5, self.G_test_1_A1_6, self.G_test_1_A1_7, self.G_test_1_A1_8, self.G_test_1_A2_1, self.G_test_1_A2_2, \
                             self.G_test_1_A3_1, self.G_test_1_A3_2, self.G_test_1_A3_3, self.G_test_1_A3_4, self.G_test_1_A3_5, \
                             self.G_test_1_A3_6, self.G_test_1_A3_7, self.G_test_1_A3_8, self.G_test_1_A3_9, self.G_test_1_A3_10, \
                             self.G_test_1_A4_1, self.G_test_1_A4_2, self.G_test_1_A4_3, self.G_test_1_A4_4, self.G_test_1_A4_5, \
                             self.G_test_1_A5_1, self.G_test_1_A5_2], 
                feed_dict={
                self.x_1 : Host_test[i*self.batch_size:(i+1)*self.batch_size].astype(np.float32),
                self.z_1 : WM_data.astype(np.float32)
            })
    
            # test generator network
            G_test_1_A6_1, G_test_1_A6_2, G_test_1_A6_3, G_test_1_A6_4, G_test_1_A6_5, G_test_1_A6_6, \
            G_test_1_A7_1, G_test_1_A7_2, G_test_1_A7_3, G_test_1_A7_4, \
            G_test_1_A7_5, G_test_1_A7_6, G_test_1_A7_7, G_test_1_A7_8, G_test_1_A7_9, \
            G_test_1_A8_1, G_test_1_A8_2, G_test_1_A8_3, G_test_1_A8_4, G_test_1_A8_5, G_test_1_A8_6, \
            G_test_1_A13_1, G_test_1_A13_2, G_test_1_A13_3, G_test_1_A13_4, G_test_1_A13_5, G_test_1_A13_6, \
            G_test_1_A13_7, G_test_1_A13_8, G_test_1_A13_9, G_test_1_A13_10, \
            G_test_1_A14_1, G_test_1_A14_2, G_test_1_A14_3, G_test_1_A14_4, G_test_1_A14_5, G_test_1_A14_6, \
            G_test_1_A14_7, G_test_1_A14_8, G_test_1_A14_9, G_test_1_A14_10, \
            G_test_1_A15_1, G_test_1_A15_2, G_test_1_A15_3, G_test_1_A15_4, G_test_1_A15_5, \
            G_test_1_A15_6, G_test_1_A15_7, G_test_1_A15_8, G_test_1_A15_9, \
            G_test_1_A16_1, G_test_1_A16_2, G_test_1_A16_3 \
                = self.sess.run([self.G_test_1_A6_1, self.G_test_1_A6_2, self.G_test_1_A6_3, self.G_test_1_A6_4, self.G_test_1_A6_5, self.G_test_1_A6_6,\
                                 self.G_test_1_A7_1, self.G_test_1_A7_2, self.G_test_1_A7_3, self.G_test_1_A7_4, self.G_test_1_A7_5, \
                                 self.G_test_1_A7_6, self.G_test_1_A7_7, self.G_test_1_A7_8, self.G_test_1_A7_9, \
                                 self.G_test_1_A8_1, self.G_test_1_A8_2, self.G_test_1_A8_3, self.G_test_1_A8_4, self.G_test_1_A8_5, self.G_test_1_A8_6, \
                                 self.G_test_1_A13_1, self.G_test_1_A13_2, self.G_test_1_A13_3, self.G_test_1_A13_4, self.G_test_1_A13_5, \
                                 self.G_test_1_A13_6, self.G_test_1_A13_7, self.G_test_1_A13_8, self.G_test_1_A13_9, self.G_test_1_A13_10,\
                                 self.G_test_1_A14_1, self.G_test_1_A14_2, self.G_test_1_A14_3, self.G_test_1_A14_4, self.G_test_1_A14_5, \
                                 self.G_test_1_A14_6, self.G_test_1_A14_7, self.G_test_1_A14_8, self.G_test_1_A14_9, self.G_test_1_A14_10,\
                                 self.G_test_1_A15_1, self.G_test_1_A15_2, self.G_test_1_A15_3, self.G_test_1_A15_4, self.G_test_1_A15_5, \
                                 self.G_test_1_A15_6, self.G_test_1_A15_7, self.G_test_1_A15_8, self.G_test_1_A15_9, \
                                 self.G_test_1_A16_1, self.G_test_1_A16_2, self.G_test_1_A16_3], 
                feed_dict={
                self.x_1 : Host_test[i*self.batch_size:(i+1)*self.batch_size].astype(np.float32),
                self.z_1 : WM_data.astype(np.float32)
            })
    
            G_test_2_,     G_test_2_A1_1, G_test_2_A1_2, G_test_2_A1_3, G_test_2_A1_4, \
            G_test_2_A1_5, G_test_2_A1_6, G_test_2_A1_7, G_test_2_A1_8, G_test_2_A2_1, G_test_2_A2_2, \
            G_test_2_A3_1, G_test_2_A3_2, G_test_2_A3_3, G_test_2_A3_4, G_test_2_A3_5, \
            G_test_2_A3_6, G_test_2_A3_7, G_test_2_A3_8, G_test_2_A3_9, G_test_2_A3_10, \
            G_test_2_A4_1, G_test_2_A4_2, G_test_2_A4_3, G_test_2_A4_4, G_test_2_A4_5, \
            G_test_2_A5_1, G_test_2_A5_2 \
            = self.sess.run([self.G_test_2_,     self.G_test_2_A1_1, self.G_test_2_A1_2, self.G_test_2_A1_3, self.G_test_2_A1_4, \
                             self.G_test_2_A1_5, self.G_test_2_A1_6, self.G_test_2_A1_7, self.G_test_2_A1_8, self.G_test_2_A2_1, self.G_test_2_A2_2, \
                             self.G_test_2_A3_1, self.G_test_2_A3_2, self.G_test_2_A3_3, self.G_test_2_A3_4, self.G_test_2_A3_5, \
                             self.G_test_2_A3_6, self.G_test_2_A3_7, self.G_test_2_A3_8, self.G_test_2_A3_9, self.G_test_2_A3_10, \
                             self.G_test_2_A4_1, self.G_test_2_A4_2, self.G_test_2_A4_3, self.G_test_2_A4_4, self.G_test_2_A4_5, \
                             self.G_test_2_A5_1, self.G_test_2_A5_2], 
                feed_dict={
                self.x_1 : Host_test[i*self.batch_size:(i+1)*self.batch_size].astype(np.float32),
                self.z_1 : WM_data.astype(np.float32)
            })
    
            # test generator network
            G_test_2_A6_1, G_test_2_A6_2, G_test_2_A6_3, G_test_2_A6_4, G_test_2_A6_5, G_test_2_A6_6, \
            G_test_2_A7_1, G_test_2_A7_2, G_test_2_A7_3, G_test_2_A7_4, \
            G_test_2_A7_5, G_test_2_A7_6, G_test_2_A7_7, G_test_2_A7_8, G_test_2_A7_9, \
            G_test_2_A8_1, G_test_2_A8_2, G_test_2_A8_3, G_test_2_A8_4, G_test_2_A8_5, G_test_2_A8_6, \
            G_test_2_A13_1, G_test_2_A13_2, G_test_2_A13_3, G_test_2_A13_4, G_test_2_A13_5, G_test_2_A13_6, \
            G_test_2_A13_7, G_test_2_A13_8, G_test_2_A13_9, G_test_2_A13_10, \
            G_test_2_A14_1, G_test_2_A14_2, G_test_2_A14_3, G_test_2_A14_4, G_test_2_A14_5, G_test_2_A14_6, \
            G_test_2_A14_7, G_test_2_A14_8, G_test_2_A14_9, G_test_2_A14_10, \
            G_test_2_A15_1, G_test_2_A15_2, G_test_2_A15_3, G_test_2_A15_4, G_test_2_A15_5, \
            G_test_2_A15_6, G_test_2_A15_7, G_test_2_A15_8, G_test_2_A15_9, \
            G_test_2_A16_1, G_test_2_A16_2, G_test_2_A16_3 \
                = self.sess.run([self.G_test_2_A6_1, self.G_test_2_A6_2, self.G_test_2_A6_3, self.G_test_2_A6_4, self.G_test_2_A6_5, self.G_test_2_A6_6,\
                                 self.G_test_2_A7_1, self.G_test_2_A7_2, self.G_test_2_A7_3, self.G_test_2_A7_4, self.G_test_2_A7_5, \
                                 self.G_test_2_A7_6, self.G_test_2_A7_7, self.G_test_2_A7_8, self.G_test_2_A7_9, \
                                 self.G_test_2_A8_1, self.G_test_2_A8_2, self.G_test_2_A8_3, self.G_test_2_A8_4, self.G_test_2_A8_5, self.G_test_2_A8_6, \
                                 self.G_test_2_A13_1, self.G_test_2_A13_2, self.G_test_2_A13_3, self.G_test_2_A13_4, self.G_test_2_A13_5, \
                                 self.G_test_2_A13_6, self.G_test_2_A13_7, self.G_test_2_A13_8, self.G_test_2_A13_9, self.G_test_2_A13_10,\
                                 self.G_test_2_A14_1, self.G_test_2_A14_2, self.G_test_2_A14_3, self.G_test_2_A14_4, self.G_test_2_A14_5, \
                                 self.G_test_2_A14_6, self.G_test_2_A14_7, self.G_test_2_A14_8, self.G_test_2_A14_9, self.G_test_2_A14_10,\
                                 self.G_test_2_A15_1, self.G_test_2_A15_2, self.G_test_2_A15_3, self.G_test_2_A15_4, self.G_test_2_A15_5, \
                                 self.G_test_2_A15_6, self.G_test_2_A15_7, self.G_test_2_A15_8, self.G_test_2_A15_9, \
                                 self.G_test_2_A16_1, self.G_test_2_A16_2, self.G_test_2_A16_3], 
                feed_dict={
                self.x_1 : Host_test[i*self.batch_size:(i+1)*self.batch_size].astype(np.float32),
                self.z_1 : WM_data.astype(np.float32)
            })

            Final_Result[i] = np.mean(PSNR_WMed_t_),np.mean(CSNR_WMed_t_), \
                np.mean(NC_WM_t_),      np.mean(NC_WM_t_A1_1),  np.mean(NC_WM_t_A1_2),  np.mean(NC_WM_t_A1_3),  np.mean(NC_WM_t_A1_4), \
                np.mean(NC_WM_t_A1_5),  np.mean(NC_WM_t_A1_6),  np.mean(NC_WM_t_A1_7),  np.mean(NC_WM_t_A1_8),  np.mean(NC_WM_t_A2_1), np.mean(NC_WM_t_A2_2), \
                np.mean(NC_WM_t_A16_1), np.mean(NC_WM_t_A16_2), np.mean(NC_WM_t_A16_3), \
                np.mean(NC_WM_t_A3_1),  np.mean(NC_WM_t_A3_2),  np.mean(NC_WM_t_A3_3),  np.mean(NC_WM_t_A3_4),  np.mean(NC_WM_t_A3_5), \
                np.mean(NC_WM_t_A3_6),  np.mean(NC_WM_t_A3_7),  np.mean(NC_WM_t_A3_8),  np.mean(NC_WM_t_A3_9),  np.mean(NC_WM_t_A3_10), \
                np.mean(NC_WM_t_A4_1),  np.mean(NC_WM_t_A4_2),  np.mean(NC_WM_t_A4_3),  np.mean(NC_WM_t_A4_4),  np.mean(NC_WM_t_A4_5), \
                np.mean(NC_WM_t_A5_1),  np.mean(NC_WM_t_A5_2), \
                np.mean(NC_WM_t_A6_1),  np.mean(NC_WM_t_A6_2),  np.mean(NC_WM_t_A6_3),  np.mean(NC_WM_t_A6_4),  np.mean(NC_WM_t_A6_5),  np.mean(NC_WM_t_A6_6),\
                np.mean(NC_WM_t_A7_1),  np.mean(NC_WM_t_A7_2),  np.mean(NC_WM_t_A7_3),  np.mean(NC_WM_t_A7_4), \
                np.mean(NC_WM_t_A7_5),  np.mean(NC_WM_t_A7_6),  np.mean(NC_WM_t_A7_7),  np.mean(NC_WM_t_A7_8),  np.mean(NC_WM_t_A7_9), \
                np.mean(NC_WM_t_A8_1),  np.mean(NC_WM_t_A8_2),  np.mean(NC_WM_t_A8_3),  np.mean(NC_WM_t_A8_4),  np.mean(NC_WM_t_A8_5),  np.mean(NC_WM_t_A8_6),\
                np.mean(NC_WM_t_A13_1), np.mean(NC_WM_t_A13_2), np.mean(NC_WM_t_A13_3), np.mean(NC_WM_t_A13_4), np.mean(NC_WM_t_A13_5), np.mean(NC_WM_t_A13_6),\
                np.mean(NC_WM_t_A13_7), np.mean(NC_WM_t_A13_8), np.mean(NC_WM_t_A13_9), np.mean(NC_WM_t_A13_10), \
                np.mean(NC_WM_t_A14_1), np.mean(NC_WM_t_A14_2), np.mean(NC_WM_t_A14_3), np.mean(NC_WM_t_A14_4), np.mean(NC_WM_t_A14_5), np.mean(NC_WM_t_A14_6), \
                np.mean(NC_WM_t_A14_7), np.mean(NC_WM_t_A14_8), np.mean(NC_WM_t_A14_9), np.mean(NC_WM_t_A14_10), \
                np.mean(NC_WM_t_A15_1), np.mean(NC_WM_t_A15_2), np.mean(NC_WM_t_A15_3), np.mean(NC_WM_t_A15_4), np.mean(NC_WM_t_A15_5), np.mean(NC_WM_t_A15_6), \
                np.mean(NC_WM_t_A15_7), np.mean(NC_WM_t_A15_8), np.mean(NC_WM_t_A15_9)
            
            print('check!!!!!!!!!!!!!!:', WMed_.shape)            
            For_Recon.append(WMed_)

            WMed = np.clip((WMed_+1)/2, 0, 1)
            Host_test_ = np.clip((Host_test+1)/2, 0, 1)
            g_test_2 = np.round(np.clip((g_test_2 +1)/2, 0, 1))
            WM_data_ = np.round(np.clip((WM_data +1)/2, 0, 1))
            counter += 1

            G_test_1_    = G_test_1_.reshape(-1, self.height,self.width)
            G_test_1_A1_1  = G_test_1_A1_1.reshape(-1, self.height,self.width)
            G_test_1_A1_2  = G_test_1_A1_2.reshape(-1, self.height,self.width)
            G_test_1_A1_3  = G_test_1_A1_3.reshape(-1, self.height,self.width)
            G_test_1_A1_4  = G_test_1_A1_4.reshape(-1, self.height,self.width)
            G_test_1_A1_5  = G_test_1_A1_5.reshape(-1, self.height,self.width)
            G_test_1_A1_6  = G_test_1_A1_6.reshape(-1, self.height,self.width)
            G_test_1_A1_7  = G_test_1_A1_7.reshape(-1, self.height,self.width)
            G_test_1_A1_8  = G_test_1_A1_8.reshape(-1, self.height,self.width)
            G_test_1_A2_1  = G_test_1_A2_1.reshape(-1, self.height,self.width)
            G_test_1_A2_2  = G_test_1_A2_2.reshape(-1, self.height,self.width)
            G_test_1_A3_1  = G_test_1_A3_1.reshape(-1, self.height,self.width)
            G_test_1_A3_2  = G_test_1_A3_2.reshape(-1, self.height,self.width)
            G_test_1_A3_3  = G_test_1_A3_3.reshape(-1, self.height,self.width)
            G_test_1_A3_4  = G_test_1_A3_4.reshape(-1, self.height,self.width)
            G_test_1_A3_5  = G_test_1_A3_5.reshape(-1, self.height,self.width)
            G_test_1_A3_6  = G_test_1_A3_6.reshape(-1, self.height,self.width)
            G_test_1_A3_7  = G_test_1_A3_7.reshape(-1, self.height,self.width)
            G_test_1_A3_8  = G_test_1_A3_8.reshape(-1, self.height,self.width)
            G_test_1_A3_9  = G_test_1_A3_9.reshape(-1, self.height,self.width)
            G_test_1_A3_10  = G_test_1_A3_10.reshape(-1, self.height,self.width)
            G_test_1_A4_1  = G_test_1_A4_1.reshape(-1, self.height,self.width)
            G_test_1_A4_2  = G_test_1_A4_2.reshape(-1, self.height,self.width)
            G_test_1_A4_3  = G_test_1_A4_3.reshape(-1, self.height,self.width)
            G_test_1_A4_4  = G_test_1_A4_4.reshape(-1, self.height,self.width)
            G_test_1_A4_5  = G_test_1_A4_5.reshape(-1, self.height,self.width)
            G_test_1_A5_1  = G_test_1_A5_1.reshape(-1, self.height,self.width)
            G_test_1_A5_2  = G_test_1_A5_2.reshape(-1, self.height,self.width)
            G_test_1_A6_1  = G_test_1_A6_1.reshape(-1, self.height,self.width)
            G_test_1_A6_2  = G_test_1_A6_2.reshape(-1, self.height,self.width)
            G_test_1_A6_3  = G_test_1_A6_3.reshape(-1, self.height,self.width)
            G_test_1_A6_4  = G_test_1_A6_4.reshape(-1, self.height,self.width)
            G_test_1_A6_5  = G_test_1_A6_5.reshape(-1, self.height,self.width)
            G_test_1_A6_6  = G_test_1_A6_6.reshape(-1, self.height,self.width)
            G_test_1_A7_1  = G_test_1_A7_1.reshape(-1, self.height,self.width)
            G_test_1_A7_2  = G_test_1_A7_2.reshape(-1, self.height,self.width)
            G_test_1_A7_3  = G_test_1_A7_3.reshape(-1, self.height,self.width)
            G_test_1_A7_4  = G_test_1_A7_4.reshape(-1, self.height,self.width)
            G_test_1_A7_5  = G_test_1_A7_5.reshape(-1, self.height,self.width)
            G_test_1_A7_6  = G_test_1_A7_6.reshape(-1, self.height,self.width)
            G_test_1_A7_7  = G_test_1_A7_7.reshape(-1, self.height,self.width)
            G_test_1_A7_8  = G_test_1_A7_8.reshape(-1, self.height,self.width)
            G_test_1_A7_9  = G_test_1_A7_9.reshape(-1, self.height,self.width)
            G_test_1_A8_1  = G_test_1_A8_1.reshape(-1, self.height,self.width)
            G_test_1_A8_2  = G_test_1_A8_2.reshape(-1, self.height,self.width)
            G_test_1_A8_3  = G_test_1_A8_3.reshape(-1, self.height,self.width)
            G_test_1_A8_4  = G_test_1_A8_4.reshape(-1, self.height,self.width)
            G_test_1_A8_5  = G_test_1_A8_5.reshape(-1, self.height,self.width)
            G_test_1_A8_6  = G_test_1_A8_6.reshape(-1, self.height,self.width)
    
            G_test_1_A13_1 = G_test_1_A13_1.reshape(-1, self.height,self.width)
            G_test_1_A13_2 = G_test_1_A13_2.reshape(-1, self.height,self.width)
            G_test_1_A13_3 = G_test_1_A13_3.reshape(-1, self.height,self.width)
            G_test_1_A13_4 = G_test_1_A13_4.reshape(-1, self.height,self.width)
            G_test_1_A13_5 = G_test_1_A13_5.reshape(-1, self.height,self.width)
            G_test_1_A13_6 = G_test_1_A13_6.reshape(-1, self.height,self.width)
            G_test_1_A13_7 = G_test_1_A13_3.reshape(-1, self.height,self.width)
            G_test_1_A13_8 = G_test_1_A13_4.reshape(-1, self.height,self.width)
            G_test_1_A13_9 = G_test_1_A13_5.reshape(-1, self.height,self.width)
            G_test_1_A13_10 = G_test_1_A13_6.reshape(-1, self.height,self.width)
            G_test_1_A14_1 = G_test_1_A14_1.reshape(-1, self.height,self.width)
            G_test_1_A14_2 = G_test_1_A14_2.reshape(-1, self.height,self.width)
            G_test_1_A14_3 = G_test_1_A14_3.reshape(-1, self.height,self.width)
            G_test_1_A14_4 = G_test_1_A14_4.reshape(-1, self.height,self.width)
            G_test_1_A14_5 = G_test_1_A14_5.reshape(-1, self.height,self.width)
            G_test_1_A14_6 = G_test_1_A14_6.reshape(-1, self.height,self.width)
            G_test_1_A14_7 = G_test_1_A14_7.reshape(-1, self.height,self.width)
            G_test_1_A14_8 = G_test_1_A14_8.reshape(-1, self.height,self.width)
            G_test_1_A14_9 = G_test_1_A14_9.reshape(-1, self.height,self.width)
            G_test_1_A14_10 = G_test_1_A14_10.reshape(-1, self.height,self.width)
            G_test_1_A15_1 = G_test_1_A15_1.reshape(-1, self.height,self.width)
            G_test_1_A15_2 = G_test_1_A15_2.reshape(-1, self.height,self.width)
            G_test_1_A15_3 = G_test_1_A15_3.reshape(-1, self.height,self.width)
            G_test_1_A15_4 = G_test_1_A15_4.reshape(-1, self.height,self.width)
            G_test_1_A15_5 = G_test_1_A15_5.reshape(-1, self.height,self.width)
            G_test_1_A15_6 = G_test_1_A15_6.reshape(-1, self.height,self.width)
            G_test_1_A15_7 = G_test_1_A15_7.reshape(-1, self.height,self.width)
            G_test_1_A15_8 = G_test_1_A15_6.reshape(-1, self.height,self.width)
            G_test_1_A15_9 = G_test_1_A15_7.reshape(-1, self.height,self.width)
            G_test_1_A16_1 = G_test_1_A16_1.reshape(-1, self.height,self.width)
            G_test_1_A16_2 = G_test_1_A16_2.reshape(-1, self.height,self.width)
            G_test_1_A16_3 = G_test_1_A16_3.reshape(-1, self.height,self.width)
            n=0
            for i in range(len(G_test_1_)):
                scipy.misc.toimage(G_test_1_[i], cmin=-1.0, cmax=1.0).save(
                    './result/' + dataset_name_ + '00_identity_' + str(i) + '.png')
                scipy.misc.toimage(WM_data_[i].reshape(round(self.height / 16), round(self.width / 16)), cmin=0,
                                   cmax=1.0).save('./result/' + dataset_name_ + '00_identity_WM' + str(i) + '.png')
            scipy.misc.toimage(G_test_1_A1_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_1(3x3).png')
            scipy.misc.toimage(G_test_1_A1_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_2(5x5).png')
            scipy.misc.toimage(G_test_1_A1_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_3(7x7).png')
            scipy.misc.toimage(G_test_1_A1_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_4(9x9).png')
            scipy.misc.toimage(G_test_1_A1_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_5(11x11).png')
            scipy.misc.toimage(G_test_1_A1_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_6(13x13).png')
            scipy.misc.toimage(G_test_1_A1_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_7(15x15).png')
            scipy.misc.toimage(G_test_1_A1_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_8(17x17).png')
            scipy.misc.toimage(G_test_1_A2_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '02_B(Avr)_1(k=5).png')
            scipy.misc.toimage(G_test_1_A2_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '02_B(Avr)_2(k=3).png')
            scipy.misc.toimage(G_test_1_A3_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_1(p=0.1).png')
            scipy.misc.toimage(G_test_1_A3_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_2(p=0.09).png')
            scipy.misc.toimage(G_test_1_A3_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_3(p=0.08).png')
            scipy.misc.toimage(G_test_1_A3_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_4(p=0.07).png')
            scipy.misc.toimage(G_test_1_A3_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_5(p=0.06).png')
            scipy.misc.toimage(G_test_1_A3_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_6(p=0.05).png')
            scipy.misc.toimage(G_test_1_A3_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_7(p=0.04).png')
            scipy.misc.toimage(G_test_1_A3_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_8(p=0.03).png')
            scipy.misc.toimage(G_test_1_A3_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_9(p=0.02).png')
            scipy.misc.toimage(G_test_1_A3_10[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_10(p=0.01).png')

            scipy.misc.toimage(G_test_1_A4_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_1(s=0.1).png')
            scipy.misc.toimage(G_test_1_A4_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_2(s=0.08).png')
            scipy.misc.toimage(G_test_1_A4_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_3(s=0.05).png')
            scipy.misc.toimage(G_test_1_A4_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_4(s=0.03).png')
            scipy.misc.toimage(G_test_1_A4_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_5(s=0.01).png')
            scipy.misc.toimage(G_test_1_A5_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '05_Sharpe_1(c=9).png')
            scipy.misc.toimage(G_test_1_A5_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '05_Sharpe_2(c=5).png')

            scipy.misc.toimage(G_test_1_A6_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_1(0.4).png')
            scipy.misc.toimage(G_test_1_A6_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_2(0.5).png')
            scipy.misc.toimage(G_test_1_A6_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_3(0.6).png')
            scipy.misc.toimage(G_test_1_A6_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_4(0.7).png')
            scipy.misc.toimage(G_test_1_A6_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_5(0.8).png')
            scipy.misc.toimage(G_test_1_A6_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_5(0.9).png')

            scipy.misc.toimage(G_test_1_A7_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_1(0.9).png')
            scipy.misc.toimage(G_test_1_A7_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_2(0.8).png')
            scipy.misc.toimage(G_test_1_A7_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_3(0.7).png')
            scipy.misc.toimage(G_test_1_A7_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_4(0.6).png')
            scipy.misc.toimage(G_test_1_A7_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_5(0.5).png')
            scipy.misc.toimage(G_test_1_A7_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_6(0.4).png')
            scipy.misc.toimage(G_test_1_A7_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_7(0.3).png')
            scipy.misc.toimage(G_test_1_A7_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_8(0.2).png')
            scipy.misc.toimage(G_test_1_A7_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_9(0.1).png')

            scipy.misc.toimage(G_test_1_A8_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_1(15).png')
            scipy.misc.toimage(G_test_1_A8_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_2(30).png')
            scipy.misc.toimage(G_test_1_A8_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_3(45).png')
            scipy.misc.toimage(G_test_1_A8_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_4(60).png')
            scipy.misc.toimage(G_test_1_A8_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_5(75).png')
            scipy.misc.toimage(G_test_1_A8_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_6(90).png')

            scipy.misc.toimage(G_test_1_A13_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_1(10).png')
            scipy.misc.toimage(G_test_1_A13_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_2(20).png')
            scipy.misc.toimage(G_test_1_A13_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_3(30).png')
            scipy.misc.toimage(G_test_1_A13_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_4(40).png')
            scipy.misc.toimage(G_test_1_A13_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_5(50).png')
            scipy.misc.toimage(G_test_1_A13_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_6(60).png')
            scipy.misc.toimage(G_test_1_A13_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_7(70).png')
            scipy.misc.toimage(G_test_1_A13_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_8(80).png')
            scipy.misc.toimage(G_test_1_A13_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_9(90).png')
            scipy.misc.toimage(G_test_1_A13_10[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_10(100).png')

            scipy.misc.toimage(G_test_1_A14_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_1(0.035).png')
            scipy.misc.toimage(G_test_1_A14_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_2(0.1).png')
            scipy.misc.toimage(G_test_1_A14_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_3(0.2).png')
            scipy.misc.toimage(G_test_1_A14_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_4(0.3).png')
            scipy.misc.toimage(G_test_1_A14_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_5(0.4).png')
            scipy.misc.toimage(G_test_1_A14_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_6(0.5).png')
            scipy.misc.toimage(G_test_1_A14_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_7(0.6).png')
            scipy.misc.toimage(G_test_1_A14_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_8(0.7).png')
            scipy.misc.toimage(G_test_1_A14_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_9(0.8).png')
            scipy.misc.toimage(G_test_1_A14_10[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_10(0.9).png')

            scipy.misc.toimage(G_test_1_A15_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_1(0.1).png')
            scipy.misc.toimage(G_test_1_A15_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_2(0.2).png')
            scipy.misc.toimage(G_test_1_A15_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_3(0.3).png')
            scipy.misc.toimage(G_test_1_A15_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_4(0.4).png')
            scipy.misc.toimage(G_test_1_A15_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_5(0.5).png')
            scipy.misc.toimage(G_test_1_A15_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_6(0.6).png')
            scipy.misc.toimage(G_test_1_A15_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_7(0.7).png')
            scipy.misc.toimage(G_test_1_A15_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_8(0.8).png')
            scipy.misc.toimage(G_test_1_A15_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_9(0.9).png')

            scipy.misc.toimage(G_test_1_A16_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '16_B(Median)_1(3x3).png')
            scipy.misc.toimage(G_test_1_A16_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '16_B(Median)_2(5x5).png')
            scipy.misc.toimage(G_test_1_A16_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '16_B(Median)_3(7x7).png')
    
            n = np.random.choice(self.batch_size)
            G_test_2_      = G_test_2_.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_1  = G_test_2_A1_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_2  = G_test_2_A1_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_3  = G_test_2_A1_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_4  = G_test_2_A1_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_5  = G_test_2_A1_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_6  = G_test_2_A1_6.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_7  = G_test_2_A1_7.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A1_8  = G_test_2_A1_8.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A2_1  = G_test_2_A2_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A2_2  = G_test_2_A2_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_1  = G_test_2_A3_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_2  = G_test_2_A3_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_3  = G_test_2_A3_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_4  = G_test_2_A3_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_5  = G_test_2_A3_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_6  = G_test_2_A3_6.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_7  = G_test_2_A3_7.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_8  = G_test_2_A3_8.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_9  = G_test_2_A3_9.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A3_10  = G_test_2_A3_10.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A4_1  = G_test_2_A4_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A4_2  = G_test_2_A4_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A4_3  = G_test_2_A4_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A4_4  = G_test_2_A4_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A4_5  = G_test_2_A4_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A5_1  = G_test_2_A5_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A5_2  = G_test_2_A5_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A6_1  = G_test_2_A6_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A6_2  = G_test_2_A6_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A6_3  = G_test_2_A6_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A6_4  = G_test_2_A6_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A6_5  = G_test_2_A6_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A6_6  = G_test_2_A6_6.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_1  = G_test_2_A7_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_2  = G_test_2_A7_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_3  = G_test_2_A7_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_4  = G_test_2_A7_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_5  = G_test_2_A7_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_6  = G_test_2_A7_6.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_7  = G_test_2_A7_7.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_8  = G_test_2_A7_8.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A7_9  = G_test_2_A7_9.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A8_1  = G_test_2_A8_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A8_2  = G_test_2_A8_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A8_3  = G_test_2_A8_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A8_4  = G_test_2_A8_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A8_5  = G_test_2_A8_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A8_6  = G_test_2_A8_6.reshape(-1, round(self.height/16),round(self.width/16))
    
            G_test_2_A13_1 = G_test_2_A13_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_2 = G_test_2_A13_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_3 = G_test_2_A13_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_4 = G_test_2_A13_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_5 = G_test_2_A13_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_6 = G_test_2_A13_6.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_7 = G_test_2_A13_7.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_8 = G_test_2_A13_8.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_9 = G_test_2_A13_9.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A13_10 = G_test_2_A13_10.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_1 = G_test_2_A14_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_2 = G_test_2_A14_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_3 = G_test_2_A14_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_4 = G_test_2_A14_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_5 = G_test_2_A14_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_6 = G_test_2_A14_6.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_7 = G_test_2_A14_7.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_8 = G_test_2_A14_8.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_9 = G_test_2_A14_9.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A14_10 = G_test_2_A14_10.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_1 = G_test_2_A15_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_2 = G_test_2_A15_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_3 = G_test_2_A15_3.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_4 = G_test_2_A15_4.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_5 = G_test_2_A15_5.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_6 = G_test_2_A15_6.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_7 = G_test_2_A15_7.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_8 = G_test_2_A15_8.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A15_9 = G_test_2_A15_9.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A16_1 = G_test_2_A16_1.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A16_2 = G_test_2_A16_2.reshape(-1, round(self.height/16),round(self.width/16))
            G_test_2_A16_3 = G_test_2_A16_3.reshape(-1, round(self.height/16),round(self.width/16))
            scipy.misc.toimage(G_test_2_[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '00_identity_WM.png')
            scipy.misc.toimage(G_test_2_A1_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_1(3x3)_WM.png')
            scipy.misc.toimage(G_test_2_A1_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_2(5x5)_WM.png')
            scipy.misc.toimage(G_test_2_A1_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_3(7x7)_WM.png')
            scipy.misc.toimage(G_test_2_A1_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_4(9x9)_WM.png')
            scipy.misc.toimage(G_test_2_A1_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_5(11x11)_WM.png')
            scipy.misc.toimage(G_test_2_A1_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_6(13x13)_WM.png')
            scipy.misc.toimage(G_test_2_A1_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_7(15x15)_WM.png')
            scipy.misc.toimage(G_test_2_A1_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '01_B(Gau)_8(17x17)_WM.png')
            scipy.misc.toimage(G_test_2_A2_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '02_B(Avr)_1(k=5)_WM.png')
            scipy.misc.toimage(G_test_2_A2_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '02_B(Avr)_2(k=3)_WM.png')
            scipy.misc.toimage(G_test_2_A3_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_1(p=0.1)_WM.png')
            scipy.misc.toimage(G_test_2_A3_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_2(p=0.09)_WM.png')
            scipy.misc.toimage(G_test_2_A3_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_3(p=0.08)_WM.png')
            scipy.misc.toimage(G_test_2_A3_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_4(p=0.07)_WM.png')
            scipy.misc.toimage(G_test_2_A3_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_5(p=0.06)_WM.png')
            scipy.misc.toimage(G_test_2_A3_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_6(p=0.05)_WM.png')
            scipy.misc.toimage(G_test_2_A3_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_7(p=0.04)_WM.png')
            scipy.misc.toimage(G_test_2_A3_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_8(p=0.03)_WM.png')
            scipy.misc.toimage(G_test_2_A3_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_9(p=0.02)_WM.png')
            scipy.misc.toimage(G_test_2_A3_10[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '03_A(S&P)_10(p=0.01)_WM.png')
            scipy.misc.toimage(G_test_2_A4_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_1(s=0.1)_WM.png')
            scipy.misc.toimage(G_test_2_A4_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_2(s=0.08)_WM.png')
            scipy.misc.toimage(G_test_2_A4_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_3(s=0.05)_WM.png')
            scipy.misc.toimage(G_test_2_A4_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_4(s=0.03)_WM.png')
            scipy.misc.toimage(G_test_2_A4_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '04_A(Gau)_5(s=0.01)_WM.png')
            scipy.misc.toimage(G_test_2_A5_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '05_Sharpe_1(c=9)_WM.png')
            scipy.misc.toimage(G_test_2_A5_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '05_Sharpe_2(c=5)_WM.png')
            scipy.misc.toimage(G_test_2_A6_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_1(0.4)_WM.png')
            scipy.misc.toimage(G_test_2_A6_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_2(0.5)_WM.png')
            scipy.misc.toimage(G_test_2_A6_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_3(0.6)_WM.png')
            scipy.misc.toimage(G_test_2_A6_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_4(0.7)_WM.png')
            scipy.misc.toimage(G_test_2_A6_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_5(0.8)_WM.png')
            scipy.misc.toimage(G_test_2_A6_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '06_G(Res)_6(0.9)_WM.png')

            scipy.misc.toimage(G_test_2_A7_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_1(0.9)_WM.png')
            scipy.misc.toimage(G_test_2_A7_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_2(0.8)_WM.png')
            scipy.misc.toimage(G_test_2_A7_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_3(0.7)_WM.png')
            scipy.misc.toimage(G_test_2_A7_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_4(0.6)_WM.png')
            scipy.misc.toimage(G_test_2_A7_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_5(0.5)_WM.png')
            scipy.misc.toimage(G_test_2_A7_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_6(0.4)_WM.png')
            scipy.misc.toimage(G_test_2_A7_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_7(0.3)_WM.png')
            scipy.misc.toimage(G_test_2_A7_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_8(0.2)_WM.png')
            scipy.misc.toimage(G_test_2_A7_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '07_Crop_out_9(0.1)_WM.png')
            scipy.misc.toimage(G_test_2_A8_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_1(15)_WM.png')
            scipy.misc.toimage(G_test_2_A8_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_2(30)_WM.png')
            scipy.misc.toimage(G_test_2_A8_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_3(45)_WM.png')
            scipy.misc.toimage(G_test_2_A8_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_4(60)_WM.png')
            scipy.misc.toimage(G_test_2_A8_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_5(75)_WM.png')
            scipy.misc.toimage(G_test_2_A8_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '08_G(Rot)_6(90)_WM.png')
            scipy.misc.toimage(G_test_2_A13_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_1(10)_WM.png')
            scipy.misc.toimage(G_test_2_A13_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_2(20)_WM.png')
            scipy.misc.toimage(G_test_2_A13_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_3(30)_WM.png')
            scipy.misc.toimage(G_test_2_A13_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_4(40)_WM.png')
            scipy.misc.toimage(G_test_2_A13_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_5(50)_WM.png')
            scipy.misc.toimage(G_test_2_A13_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_6(60)_WM.png')
            scipy.misc.toimage(G_test_2_A13_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_7(70)_WM.png')
            scipy.misc.toimage(G_test_2_A13_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_8(80)_WM.png')
            scipy.misc.toimage(G_test_2_A13_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_9(90)_WM.png')
            scipy.misc.toimage(G_test_2_A13_10[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '13_JPEG_10(100)_WM.png')
            scipy.misc.toimage(G_test_2_A14_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_1(0.035)_WM.png')
            scipy.misc.toimage(G_test_2_A14_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_2(0.1)_WM.png')
            scipy.misc.toimage(G_test_2_A14_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_3(0.2)_WM.png')
            scipy.misc.toimage(G_test_2_A14_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_4(0.3)_WM.png')
            scipy.misc.toimage(G_test_2_A14_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_5(0.4)_WM.png')
            scipy.misc.toimage(G_test_2_A14_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_6(0.5)_WM.png')
            scipy.misc.toimage(G_test_2_A14_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_7(0.6)_WM.png')
            scipy.misc.toimage(G_test_2_A14_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_8(0.7)_WM.png')
            scipy.misc.toimage(G_test_2_A14_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_9(0.8)_WM.png')
            scipy.misc.toimage(G_test_2_A14_10[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '14_Grid_Crop_10(0.9)_WM.png')
            scipy.misc.toimage(G_test_2_A15_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_1(0.1)_WM.png')
            scipy.misc.toimage(G_test_2_A15_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_2(0.2)_WM.png')
            scipy.misc.toimage(G_test_2_A15_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_3(0.3)_WM.png')
            scipy.misc.toimage(G_test_2_A15_4[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_4(0.4)_WM.png')
            scipy.misc.toimage(G_test_2_A15_5[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_5(0.5)_WM.png')
            scipy.misc.toimage(G_test_2_A15_6[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_6(0.6)_WM.png')
            scipy.misc.toimage(G_test_2_A15_7[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_7(0.7)_WM.png')
            scipy.misc.toimage(G_test_2_A15_8[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_8(0.8)_WM.png')
            scipy.misc.toimage(G_test_2_A15_9[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '15_Dropout_9(0.9)_WM.png')
            scipy.misc.toimage(G_test_2_A16_1[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '16_B(Median)_1(3x3)_WM.png')
            scipy.misc.toimage(G_test_2_A16_2[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '16_B(Median)_2(5x5)_WM.png')
            scipy.misc.toimage(G_test_2_A16_3[n], cmin=-1.0, cmax=1.0).save(
                './result/' + dataset_name_ + '16_B(Median)_3(7x7)_WM.png')


        Final_Result_ = np.mean(Final_Result, axis=0)
        For_Recon = np.asarray(For_Recon, dtype=np.float32)
        For_Recon = For_Recon.reshape([-1, self.height, self.width, 1])
        For_Recon = (For_Recon+1)/2 * (self.in_maxmin)+self.in_min
        # np.save('./for_reconstruction_npy/v9_holo_amplitude_network/%s.npy' %(listdir) , For_Recon)


        # log print
        print('[%s] Epoch : [%2d], time : %4.4fs, PSNR_WMed_test:%.2f, CSNR_WMed_test:%.2f, \n identi : [%.4f], \n B(Gau) : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f], (3x3, 5x5, 7x7, 9x9, 11x11, 13x13, 15x15, 17x17), \n B(Avr) : [%.4f/ %.4f], (5x5, 3x3), \n B(Median) : [%.4f/ %.4f/ %.4f], (3x3, 5x5, 7x7), \n A(S&P) : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f], (0.1, 0.09, 0.08, 0.07, 0.06, 0.05, 0.04, 0.03, 0.02, 0.01) \n A(Gau) : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f], (0.1, 0.08, 0.05, 0.03, 0.01) \n Sharpe : [%.4f/ %.4f], (9, 5), \n G(Res) : [%.4f/%.4f/ %.4f/ %.4f/ %.4f/ %.4f], (0.4, 0.5, 0.6, 0.7, 0.8, 0.9) \n G(Crp) : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f], (0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2, 0.1) \n G(Rot) : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f], (15, 30, 45, 60, 75, 90)\n JPEG__ : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f], (10, 20, 30, 40, 50, 60, 70, 80, 90, 100)\n Grid_Crop : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f], (0.035, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9)\n Dropout : [%.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f/ %.4f], (0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9)' 
            % (listdir, start_epoch-1, time.time()-start_time, Final_Result_[0], \
                Final_Result_[1],  Final_Result_[2],  Final_Result_[3], Final_Result_[4], Final_Result_[5], \
                Final_Result_[6],  Final_Result_[7],  Final_Result_[8], Final_Result_[9], Final_Result_[10], Final_Result_[11], \
                Final_Result_[12], Final_Result_[13], Final_Result_[14], \
                Final_Result_[15], Final_Result_[16], Final_Result_[17], Final_Result_[18], Final_Result_[19], \
                Final_Result_[20], Final_Result_[21], Final_Result_[22], Final_Result_[23], Final_Result_[24], \
                Final_Result_[25], Final_Result_[26], Final_Result_[27], Final_Result_[28], Final_Result_[29], \
                Final_Result_[30], Final_Result_[31],  \
                Final_Result_[32], Final_Result_[33], Final_Result_[34], Final_Result_[35], Final_Result_[36], Final_Result_[37],\
                Final_Result_[38], Final_Result_[39], Final_Result_[40], Final_Result_[41], \
                Final_Result_[42], Final_Result_[43], Final_Result_[44], Final_Result_[45], Final_Result_[46], \
                Final_Result_[47], Final_Result_[48], Final_Result_[49], Final_Result_[50], Final_Result_[51], Final_Result_[52],\
                Final_Result_[53], Final_Result_[54], Final_Result_[55], Final_Result_[56], Final_Result_[57], Final_Result_[58],\
                Final_Result_[59], Final_Result_[60], Final_Result_[61], Final_Result_[62], \
                Final_Result_[63], Final_Result_[64], Final_Result_[65], Final_Result_[66], Final_Result_[67], Final_Result_[68], \
                Final_Result_[69], Final_Result_[70], Final_Result_[71], Final_Result_[72], \
                Final_Result_[73], Final_Result_[74], Final_Result_[75], Final_Result_[76], Final_Result_[77], Final_Result_[78], \
                Final_Result_[79], Final_Result_[80], Final_Result_[81], Final_Result_[82]))

        text_list=['[%s], alpha : [%.2f]\n Epoch : [%2d], time : %4.4fs, PSNR_WMed_test: %.2f, CSNR_WMed_test:%.2f, \n identi : [ %.4f ], \n B(Gau) : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f ], (3x3, 5x5, 7x7, 9x9, 11x11, 13x13, 15x15, 17x17), \n B(Avr) : [ %.4f / %.4f ], (5x5, 3x3), \n B(Median) : [ %.4f / %.4f / %.4f ], (3x3, 5x5, 7x7), \n A(S&P) : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f/ %.4f ], (0.1, 0.09, 0.08, 0.07, 0.06, 0.05, 0.04, 0.03, 0.02, 0.01) \n A(Gau) : [ %.4f / %.4f / %.4f / %.4f / %.4f ], (0.1, 0.08, 0.05, 0.03, 0.01) \n Sharpe : [ %.4f / %.4f ], (9, 5), \n G(Res) : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f ], (0.4, 0.5, 0.6, 0.7, 0.8, 0.9) \n G(Crp) : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f ], (0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2, 0.1) \n G(Rot) : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f ], (15, 30, 45, 60, 75, 90)\n JPEG__ : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f ], (10, 20, 30, 40, 50, 60, 70, 80, 90, 100)\n Grid_Crop : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f ], (0.035, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9)\n Dropout : [ %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f / %.4f ], (0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9)' 
            % (listdir, self.alpha, start_epoch-1, time.time()-start_time, Final_Result_[0], \
                Final_Result_[1],  Final_Result_[2],  Final_Result_[3], Final_Result_[4], Final_Result_[5], \
                Final_Result_[6],  Final_Result_[7],  Final_Result_[8], Final_Result_[9], Final_Result_[10], Final_Result_[11], \
                Final_Result_[12], Final_Result_[13], Final_Result_[14], \
                Final_Result_[15], Final_Result_[16], Final_Result_[17], Final_Result_[18], Final_Result_[19], \
                Final_Result_[20], Final_Result_[21], Final_Result_[22], Final_Result_[23], Final_Result_[24], \
                Final_Result_[25], Final_Result_[26], Final_Result_[27], Final_Result_[28], Final_Result_[29], \
                Final_Result_[30], Final_Result_[31],  \
                Final_Result_[32], Final_Result_[33], Final_Result_[34], Final_Result_[35], Final_Result_[36], Final_Result_[37],\
                Final_Result_[38], Final_Result_[39], Final_Result_[40], Final_Result_[41], \
                Final_Result_[42], Final_Result_[43], Final_Result_[44], Final_Result_[45], Final_Result_[46], \
                Final_Result_[47], Final_Result_[48], Final_Result_[49], Final_Result_[50], Final_Result_[51], Final_Result_[52],\
                Final_Result_[53], Final_Result_[54], Final_Result_[55], Final_Result_[56], Final_Result_[57], Final_Result_[58],\
                Final_Result_[59], Final_Result_[60], Final_Result_[61], Final_Result_[62], \
                Final_Result_[63], Final_Result_[64], Final_Result_[65], Final_Result_[66], Final_Result_[67], Final_Result_[68], \
                Final_Result_[69], Final_Result_[70], Final_Result_[71], Final_Result_[72], \
                Final_Result_[73], Final_Result_[74], Final_Result_[75], Final_Result_[76], Final_Result_[77], Final_Result_[78], \
                Final_Result_[79], Final_Result_[80], Final_Result_[81], Final_Result_[82])]

        
        f = open('./result/'+dataset_name+'/test_result_for_reconstruction_128x128_s1.txt', 'a')
        f.write('\n')
        f.writelines('\n'.join(text_list))
        f.close()

        arr = Final_Result_
        return arr


        
    def load_checkpoint(self, checkpoint_dir, dataset_name, ep):
        import re
        checkpoint_dir = os.path.join(checkpoint_dir, dataset_name)

        #ckpt = tf.train.get_checkpoint_state(checkpoint_dir)
        #if ckpt and ckpt.model_checkpoint_path:
        #    ckpt_name = os.path.basename(ckpt.model_checkpoint_path)
        ckpt_name = dataset_name+'.model-'+str(53050)
        self.saver.restore(self.sess, os.path.join(checkpoint_dir, ckpt_name))
        counter = int(next(re.finditer("(\d+)(?!.*\d)", ckpt_name)).group(0))
        print(" [*] Success to read {}".format(ckpt_name))
        return True, counter

  

if __name__ == '__main__' :

    width = 128    
    height = 128      

    path = "v9_holo_amplitude"
    ep = 3999
    
    alpha = 1.0
    # data = "01_ERC_hologram_test"
    print(path, ep)

    # PATH = './test_data_reim/'
    # #PATH = '../../01_Hologram_h5/test_data/batch_16/'
    # listdir = os.listdir(PATH)
    minn = 2.249197042683758e-08
    maxmin = 4.130666773016852
    i = 1000
    listdirr = "test_65000"
    model = WM_NN(width=width, height=height, kernel1=3, kernel2=3, batch_size=100,
                          d1_lr=0.0001, g1_lr=0.0001, gan2_lr=0.00001, beta1=0.5, beta2=0.999,
                          lambda1=45.0, lambda2=0.2, lambda3=20.0, alpha=alpha,in_maxmin=maxmin,in_min=minn)
    model.build_model()
    model.build_model_test()
    wb = Workbook()
    ws = wb.create_sheet(title='ERC_v2')
    arr = model.test(path, ep=ep+1, listdir=listdirr)
    for row in range(len(arr)+1):
        if row == 0 :
            ws.cell(row=row+1, column=1+i, value=listdirr)
        else :
            ws.cell(row=row+1, column=1+i, value=arr[row-1])

    wb.save('./for_reconstruction_npy/'+path+'/exel/bt16.xlsx')